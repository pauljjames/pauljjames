"""Writing the spreadsheets this tool hands out.

The counterpart to importer.py, which reads them. The header row is built from
importer.TIMETABLE_COLUMNS, so a file written here is always one that module can
read back: a template whose columns the parser rejects would be worse than no
template at all.

Two of the columns an export carries are not part of the format. A course name
and who is teaching are there to be read, and the parser ignores them on the way
back in. Staffing is decided in the Planner, where every assignment is checked
against everything else that person teaches; a spreadsheet is not going to be a
second way in past that check.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import importer

READ_ONLY = ("Course Name", "Assigned")

WIDTHS = {
    "Course Code": 13, "Section": 12, "Day": 12, "Start": 8, "End": 8,
    "Weeks": 20, "Course Name": 30, "Assigned": 46,
}

FILL_IN = (Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="1B2432"))
LEAVE_BE = (Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="7A8598"))

EXAMPLES = [
    ("133150", "SHOW-A", "Tuesday", "14:00", "17:00", "1-12"),
    ("133154", "LEC", "Monday", "09:00", "10:00", "7-9"),
    ("133154", "WS-A", "Monday", "10:00", "12:00", "1-6, 8, 10-12"),
]

NOTES = [
    "One row per class. A course with four sections is four rows.",
    "",
    "Weeks     1-12, or 1-6, 8, 10-12. A break is a gap, not an extra week.",
    "Day       Monday, or Mon. Times as 14:00, 2pm or 1430.",
    "Section   whatever your timetable calls it: LEC, WS-A, SHOW-B.",
    "",
    "Course names are not read from this file. They come from the course",
    "catalogue on the Courses page, so a course is named in one place.",
    "",
    "Who teaches a class is not read from this file either. That is decided",
    "on the Planner, where every assignment is checked against everything",
    "else that person already teaches.",
]


def _header(sheet, columns: tuple[str, ...]) -> None:
    sheet.append(list(columns))
    for index, name in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.font, cell.fill = LEAVE_BE if name in READ_ONLY else FILL_IN
        cell.alignment = Alignment(horizontal="left")
        sheet.column_dimensions[get_column_letter(index)].width = WIDTHS.get(name, 16)
    sheet.freeze_panes = "A2"


def _bytes(book: Workbook) -> bytes:
    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def timetable_template() -> bytes:
    """A blank timetable to fill in, and a second sheet showing how."""
    book = Workbook()

    sheet = book.active
    sheet.title = "Timetable"
    _header(sheet, importer.TIMETABLE_COLUMNS)

    # Worked rows live on their own sheet so nothing illustrative can be
    # imported by accident along with the real thing.
    example = book.create_sheet("Example")
    _header(example, importer.TIMETABLE_COLUMNS)
    for row in EXAMPLES:
        example.append(list(row))

    example.append([])
    for line in NOTES:
        example.append([line])

    return _bytes(book)


def timetable_export(rows: list[dict]) -> bytes:
    """The current timetable, editable in Excel and importable again.

    Each row wants course_code, section, day, start, end, weeks, and optionally
    course_name and assigned, the latter a list of {name, weeks}.
    """
    book = Workbook()
    sheet = book.active
    sheet.title = "Timetable"
    _header(sheet, importer.TIMETABLE_COLUMNS + READ_ONLY)

    for row in rows:
        sheet.append([
            row.get("course_code", ""),
            row.get("section", ""),
            row.get("day", ""),
            row.get("start", ""),
            row.get("end", ""),
            importer.format_weeks(row.get("weeks") or []),
            row.get("course_name", ""),
            "; ".join(
                f"{span['name']} {importer.format_weeks(span['weeks'])}"
                for span in row.get("assigned") or []
            ),
        ])

    return _bytes(book)


# ------------------------------------------------------------ staff calendars

SLOT = 30                       # minutes per grid row
CLASS_FILL = PatternFill("solid", fgColor="E8EEF7")
TIME_FONT = Font(color="7A8598", size=9)
TITLE_FONT = Font(bold=True, size=14)
EDGE = Border(*[Side(style="thin", color="C9D2E0")] * 4)


def _hhmm(minutes: int) -> str:
    return f"{minutes // 60:02d}:{minutes % 60:02d}"


def _minutes(clock: str) -> int:
    hours, _, mins = clock.partition(":")
    return int(hours) * 60 + int(mins)


def _calendar(sheet, classes: list[dict], at_row: int) -> int:
    """Lay a week out as a calendar: half hours down, days across.

    Returns the row after the grid. The grid spans only the hours the person
    actually teaches, because a day that starts at nine and ends at five is
    sixteen rows of nothing on either side of the part worth reading.
    """
    timed = [c for c in classes if c.get("day") and c.get("start") and c.get("end")]
    if not timed:
        sheet.cell(row=at_row, column=1, value="Nothing timetabled.").font = TIME_FONT
        return at_row + 2

    days = [d for d in importer.WEEKDAYS if any(c["day"] == d for c in timed)]
    first = min(_minutes(c["start"]) for c in timed) // SLOT * SLOT
    last = -(-max(_minutes(c["end"]) for c in timed) // SLOT) * SLOT

    header = at_row
    sheet.cell(row=header, column=1, value="Time")
    for index, day in enumerate(days, start=2):
        sheet.cell(row=header, column=index, value=day)
    for index in range(1, len(days) + 2):
        cell = sheet.cell(row=header, column=index)
        cell.font, cell.fill = FILL_IN
        cell.alignment = Alignment(horizontal="left")
        sheet.column_dimensions[get_column_letter(index)].width = 9 if index == 1 else 24

    slots = list(range(first, last, SLOT))
    for offset, minute in enumerate(slots):
        cell = sheet.cell(row=header + 1 + offset, column=1, value=_hhmm(minute))
        cell.font = TIME_FONT

    for c in timed:
        column = days.index(c["day"]) + 2
        top = header + 1 + (_minutes(c["start"]) - first) // SLOT
        deep = max(1, (_minutes(c["end"]) - _minutes(c["start"])) // SLOT)
        label = f"{c['code']} {c['section']}"
        if c.get("name"):
            label += f"\n{c['name']}"
        label += f"\n{c['start']}–{c['end']}"
        if not c.get("runs", True):
            label += "\n(cancelled)"

        cell = sheet.cell(row=top, column=column, value=label)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.fill = CLASS_FILL
        cell.border = EDGE
        if deep > 1:
            sheet.merge_cells(start_row=top, start_column=column,
                              end_row=top + deep - 1, end_column=column)

    return header + len(slots) + 2


def staff_calendars(people: list[dict], weeks: list[int]) -> bytes:
    """Everyone's teaching: a summary of hours, then a calendar each.

    A person's sheet is the week they repeat, drawn as a week rather than
    listed as rows, followed by the weeks that depart from it. That is the same
    shape the Staff page shows, and for the same reason: a semester is mostly
    one week over and over, and printing twelve near-identical weeks is how a
    useful document becomes one nobody reads.

    Each person wants name, email, target_minutes, usual_weeks, usual (a list
    of {code, section, name, day, start, end, runs}), departures (a list of
    {weeks, lines}) and load ({week: minutes}).
    """
    book = Workbook()

    summary = book.active
    summary.title = "Summary"
    _header(summary, ("Staff",) + tuple(f"Wk {w}" for w in weeks) + ("Average",))
    summary.column_dimensions["A"].width = 24
    for index in range(2, len(weeks) + 3):
        summary.column_dimensions[get_column_letter(index)].width = 7
    for person in people:
        load = person.get("load") or {}
        hours = [round(load.get(w, 0) / 60, 1) for w in weeks]
        average = round(sum(hours) / len(hours), 1) if hours else 0
        summary.append([person["name"], *hours, average])

    used: set[str] = set()
    for person in people:
        sheet = book.create_sheet(_tab(person["name"], used))

        sheet.cell(row=1, column=1, value=person["name"]).font = TITLE_FONT
        about = [person.get("email") or ""]
        if person.get("target_minutes"):
            about.append(f"target {round(person['target_minutes'] / 60, 1)} h a week")
        sheet.cell(row=2, column=1, value=" · ".join(p for p in about if p))

        usual_weeks = person.get("usual_weeks") or []
        sheet.cell(row=4, column=1, value=(
            f"The week they repeat — weeks {importer.format_weeks(usual_weeks)}"
            if usual_weeks else "No repeating week."
        )).font = Font(bold=True)

        row = _calendar(sheet, person.get("usual") or [], 5)

        departures = person.get("departures") or []
        if departures:
            sheet.cell(row=row, column=1,
                       value="Weeks that depart from this").font = Font(bold=True)
            row += 1
            for departure in departures:
                sheet.cell(row=row, column=1,
                           value=importer.format_weeks(departure["weeks"]))
                sheet.cell(row=row, column=2, value="; ".join(departure["lines"]))
                row += 1

    return _bytes(book)


def _tab(name: str, used: set[str]) -> str:
    """A sheet name Excel will accept, and has not had before."""
    clean = "".join(" " if ch in "[]:*?/\\" else ch for ch in name).strip()[:31]
    clean = clean or "Somebody"
    candidate, n = clean, 2
    while candidate.lower() in used:
        suffix = f" {n}"
        candidate = clean[: 31 - len(suffix)] + suffix
        n += 1
    used.add(candidate.lower())
    return candidate
