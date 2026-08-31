"""Reading an externally published timetable out of a spreadsheet.

The timetable is not this tool's document. It arrives as a file, is corrected
and reissued, and has to be read as it comes rather than as we would like it.
So this module is forgiving about shape and unforgiving about meaning: it will
accept a header called "Course", "course code" or "Code", but it will not guess
what "Thur 9" means. Anything it cannot read becomes a numbered complaint
against a row, and the caller shows the whole list before writing anything.

Pure parsing. No database, no HTTP.
"""

from __future__ import annotations

import csv
import io
import re

from engine import WEEKDAYS

# Every spelling we will answer to, mapped to the field we mean.
HEADERS = {
    "course_code": ("course code", "course", "code", "unit code", "module code"),
    "course_title": ("course title", "title", "course name", "name", "unit", "module"),
    "section": ("section", "class", "group", "activity", "component", "type"),
    "day": ("day", "weekday", "day of week"),
    "start": ("start", "start time", "from", "begins"),
    "end": ("end", "end time", "to", "finish", "ends", "until"),
    "weeks": ("weeks", "week", "teaching weeks", "week numbers", "runs"),
}

REQUIRED = ("course_code", "section", "day", "start", "end", "weeks")

DAY_LOOKUP = {}
for _day in WEEKDAYS:
    DAY_LOOKUP[_day.lower()] = _day
    DAY_LOOKUP[_day.lower()[:3]] = _day
DAY_LOOKUP.update({"tues": "Tuesday", "thur": "Thursday", "thurs": "Thursday"})


class ImportError_(Exception):
    """The file itself could not be read at all."""


def parse(filename: str, data: bytes) -> tuple[list[dict], list[str]]:
    """Read a timetable file into rows ready for store.replace_timetable.

    Returns the rows it understood and the complaints it has. Rows with a
    complaint are left out, so a caller that ignores the complaints imports
    less rather than importing nonsense.
    """
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        table = _read_xlsx(data)
    else:
        table = _read_csv(data)

    if not table:
        raise ImportError_("The file is empty.")

    header, *body = table
    columns, header_issues = _map_headers(header)
    if header_issues:
        return [], header_issues

    rows: list[dict] = []
    issues: list[str] = []
    seen: dict[tuple[str, str], int] = {}

    for offset, raw in enumerate(body):
        line = offset + 2   # what a spreadsheet would call this row
        if not any(str(cell).strip() for cell in raw if cell is not None):
            continue

        values = {
            field: _cell(raw, index) for field, index in columns.items()
        }
        row, problems = _build(values, line)
        if problems:
            issues.extend(problems)
            continue

        key = (row["course_code"], row["section"])
        if key in seen:
            issues.append(
                f"Row {line}: {row['course_code']} {row['section']} is also on "
                f"row {seen[key]}. Two rows for one section make duplicate "
                "classes, so this row was left out."
            )
            continue
        seen[key] = line
        rows.append(row)

    rows.sort(key=lambda r: (r["course_code"], r["section"]))
    return rows, issues


# ------------------------------------------------------------ file shapes

def _read_csv(data: bytes) -> list[list]:
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = data.decode("latin-1")
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    return [row for row in csv.reader(io.StringIO(text), dialect)]


def _read_xlsx(data: bytes) -> list[list]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError:  # pragma: no cover - depends on the environment
        raise ImportError_(
            "Reading .xlsx needs openpyxl. Install it, or save the file as CSV."
        )
    book = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet = book.active
    return [list(row) for row in sheet.iter_rows(values_only=True)]


# ------------------------------------------------------------ headers

def _map_headers(header: list) -> tuple[dict[str, int], list[str]]:
    found: dict[str, int] = {}
    for index, cell in enumerate(header):
        label = _clean(cell).lower().replace("_", " ")
        for field, spellings in HEADERS.items():
            if field not in found and label in spellings:
                found[field] = index

    missing = [f for f in REQUIRED if f not in found]
    if missing:
        return {}, [
            "The file needs a header row with columns for "
            + ", ".join(m.replace("_", " ") for m in missing)
            + ". Found: "
            + (", ".join(_clean(c) for c in header if _clean(c)) or "nothing")
            + "."
        ]
    return found, []


def _cell(row: list, index: int):
    return row[index] if index < len(row) else None


def _clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


# ------------------------------------------------------------ one row

def _build(values: dict, line: int) -> tuple[dict, list[str]]:
    issues: list[str] = []

    code = _clean(values.get("course_code"))
    section = _clean(values.get("section"))
    if not code:
        issues.append(f"Row {line}: no course code.")
    if not section:
        issues.append(f"Row {line}: no section.")

    day = _day(values.get("day"))
    if day is None:
        issues.append(f"Row {line}: {_clean(values.get('day'))!r} is not a day of the week.")

    start = _time(values.get("start"))
    end = _time(values.get("end"))
    if start is None:
        issues.append(f"Row {line}: {_clean(values.get('start'))!r} is not a time.")
    if end is None:
        issues.append(f"Row {line}: {_clean(values.get('end'))!r} is not a time.")
    if start and end and end <= start:
        issues.append(f"Row {line}: ends at or before it starts.")

    weeks, week_issue = _weeks(values.get("weeks"))
    if week_issue:
        issues.append(f"Row {line}: {week_issue}")

    if issues:
        return {}, issues

    return {
        "course_code": code,
        "course_title": _clean(values.get("course_title")),
        "section": section,
        "day": day,
        "start": start,
        "end": end,
        "weeks": weeks,
    }, []


def _day(value) -> str | None:
    text = _clean(value).lower().rstrip(".")
    return DAY_LOOKUP.get(text)


def _time(value) -> str | None:
    """Accept 9:00, 09:00, 9.00, 0900, 9am, and what a spreadsheet gives us."""
    if value is None:
        return None
    if hasattr(value, "hour") and hasattr(value, "minute"):
        return f"{value.hour:02d}:{value.minute:02d}"

    text = _clean(value).lower().replace(" ", "")
    if not text:
        return None

    suffix = ""
    if text.endswith(("am", "pm")):
        suffix, text = text[-2:], text[:-2]

    match = re.fullmatch(r"(\d{1,2})[:.h]?(\d{2})?", text)
    if not match:
        return None
    hour = int(match.group(1))
    minute = int(match.group(2) or 0)

    if suffix == "pm" and hour < 12:
        hour += 12
    elif suffix == "am" and hour == 12:
        hour = 0

    if hour > 23 or minute > 59:
        return None
    return f"{hour:02d}:{minute:02d}"


def _weeks(value) -> tuple[list[int], str | None]:
    """Read "1-6, 8, 10-12" into the weeks it means."""
    text = _clean(value)
    # "Weeks 1-6" and "wks 1-6" mean the same as "1-6".
    text = re.sub(r"\b(weeks?|wks?)\b", " ", text, flags=re.IGNORECASE).strip()
    if not text:
        return [], "no weeks, so it would never run."

    weeks: set[int] = set()
    for part in re.split(r"[,;&]| and ", text):
        part = part.strip()
        if not part:
            continue
        span = re.fullmatch(r"(\d+)\s*(?:-|–|—|to|\.\.)\s*(\d+)", part)
        if span:
            first, last = int(span.group(1)), int(span.group(2))
            if last < first:
                return [], f"{part!r} runs backwards."
            weeks.update(range(first, last + 1))
            continue
        if part.isdigit():
            weeks.add(int(part))
            continue
        return [], f"{part!r} is not a week or a range of weeks."

    if not weeks:
        return [], "no weeks, so it would never run."
    if min(weeks) < 1:
        return [], "weeks start at 1."
    return sorted(weeks), None
